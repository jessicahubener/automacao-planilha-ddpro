import tkinter as tk
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


NOME_ABA_PADRAO = "Cadastro"
NOME_ABA_RESULTADO = "Resultado"
TITULO_STATUS = "Status da Calibracao"
COLUNAS_DATA_CANDIDATAS = [
    "proxima calibracao",
    "data da proxima calibracao",
    "data de vencimento",
]
COLUNAS_CODIGO_CANDIDATAS = ["codigo"]


def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(char for char in texto if not unicodedata.combining(char))


def verificar_status(data, hoje):
    if data is None:
        return "Data invalida"
    if isinstance(data, datetime):
        data = data.date()
    elif isinstance(data, date):
        pass
    elif hasattr(data, "date") and not isinstance(data, str):
        data = data.date()
    else:
        return "Data invalida"

    if data < hoje:
        return "Vencido"
    if data <= hoje + timedelta(days=30):
        return "Vence em breve"
    return "Em dia"


def encontrar_linha_cabecalho(ws):
    limite = min(ws.max_row, 15)
    for linha in range(1, limite + 1):
        valores = [
            normalizar_texto(ws.cell(linha, coluna).value)
            for coluna in range(1, ws.max_column + 1)
        ]
        if any(candidata in valores for candidata in COLUNAS_DATA_CANDIDATAS):
            return linha
    raise ValueError("Nao foi possivel localizar o cabecalho da planilha.")


def encontrar_coluna(ws, linha_cabecalho, candidatas):
    for coluna in range(1, ws.max_column + 1):
        valor = normalizar_texto(ws.cell(linha_cabecalho, coluna).value)
        if valor in candidatas:
            return coluna, ws.cell(linha_cabecalho, coluna).value
    raise ValueError("Coluna necessaria nao encontrada na planilha.")


def definir_arquivo_saida(arquivo_entrada):
    base = arquivo_entrada.with_name(f"{arquivo_entrada.stem}_resultado.xlsx")
    if not base.exists():
        return base

    indice = 1
    while True:
        candidato = arquivo_entrada.with_name(
            f"{arquivo_entrada.stem}_resultado_{indice}.xlsx"
        )
        if not candidato.exists():
            return candidato
        indice += 1


def definir_arquivo_saida_alternativo(arquivo_entrada):
    pasta_local = Path(__file__).resolve().parent
    base = pasta_local / f"{arquivo_entrada.stem}_resultado.xlsx"
    if not base.exists():
        return base

    indice = 1
    while True:
        candidato = pasta_local / f"{arquivo_entrada.stem}_resultado_{indice}.xlsx"
        if not candidato.exists():
            return candidato
        indice += 1


def formatar_data(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def criar_planilha_resultado(registros, arquivo_saida):
    wb = Workbook()
    ws = wb.active
    ws.title = NOME_ABA_RESULTADO
    preenchimento_vencido = PatternFill(
        fill_type="solid",
        start_color="FF0000",
        end_color="FF0000",
    )

    cabecalhos = [
        "Familia",
        "Codigo",
        "Descricao",
        "Descricao da Inspecao",
        "Setor Utilizado",
        "Unidade de Medida",
        "Data da Calibracao",
        "Proxima Calibracao",
        TITULO_STATUS,
    ]
    ws.append(cabecalhos)

    for celula in ws[1]:
        celula.font = Font(bold=True)

    for registro in registros:
        ws.append(registro)
        if registro[-1] == "Vencido":
            ws.cell(ws.max_row, 9).fill = preenchimento_vencido

    larguras = {
        "A": 16,
        "B": 16,
        "C": 55,
        "D": 35,
        "E": 28,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 22,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.freeze_panes = "A2"
    wb.save(arquivo_saida)
    wb.close()


def processar_planilha(caminho_arquivo):
    arquivo_entrada = Path(caminho_arquivo)
    if not arquivo_entrada.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {arquivo_entrada}")

    arquivo_saida = definir_arquivo_saida(arquivo_entrada)
    wb_valores = load_workbook(arquivo_entrada, data_only=True)

    if NOME_ABA_PADRAO not in wb_valores.sheetnames:
        raise ValueError(f"A aba '{NOME_ABA_PADRAO}' nao foi encontrada.")

    ws_valores = wb_valores[NOME_ABA_PADRAO]
    linha_cabecalho = encontrar_linha_cabecalho(ws_valores)
    coluna_data, nome_coluna_data = encontrar_coluna(
        ws_valores, linha_cabecalho, COLUNAS_DATA_CANDIDATAS
    )
    coluna_codigo, _ = encontrar_coluna(
        ws_valores, linha_cabecalho, COLUNAS_CODIGO_CANDIDATAS
    )

    hoje = datetime.today().date()
    contagem = {"Vencido": 0, "Vence em breve": 0, "Em dia": 0, "Data invalida": 0}
    registros = []

    for linha in range(linha_cabecalho + 1, ws_valores.max_row + 1):
        codigo = ws_valores.cell(linha, coluna_codigo).value
        data_calibracao = ws_valores.cell(linha, 3).value
        proxima_calibracao = ws_valores.cell(linha, coluna_data).value

        if codigo in (None, ""):
            continue
        if data_calibracao in (None, "") and proxima_calibracao in (None, ""):
            continue

        status = verificar_status(proxima_calibracao, hoje)
        contagem[status] = contagem.get(status, 0) + 1

        registros.append(
            [
                ws_valores.cell(linha, 1).value or "",
                codigo,
                ws_valores.cell(linha, 7).value or "",
                ws_valores.cell(linha, 8).value or "",
                ws_valores.cell(linha, 9).value or "",
                ws_valores.cell(linha, 10).value or "",
                formatar_data(data_calibracao),
                formatar_data(proxima_calibracao),
                status,
            ]
        )

    wb_valores.close()

    try:
        criar_planilha_resultado(registros, arquivo_saida)
    except PermissionError:
        arquivo_saida = definir_arquivo_saida_alternativo(arquivo_entrada)
        criar_planilha_resultado(registros, arquivo_saida)

    contagem = {chave: valor for chave, valor in contagem.items() if valor > 0}
    return arquivo_saida, contagem, nome_coluna_data


class AppCalibracao:
    def __init__(self, root):
        self.root = root
        self.root.title("Controle de Calibracao")
        self.root.geometry("680x340")
        self.root.resizable(False, False)

        self.caminho_arquivo = tk.StringVar()
        self.status_texto = tk.StringVar(
            value="Selecione a planilha para gerar um Excel organizado com os resultados."
        )
        self.montar_tela()

    def montar_tela(self):
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="Controle de Calibracao",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", pady=(0, 12))

        ttk.Label(
            frame,
            text="O arquivo novo sera um Excel com nomes, datas e status em colunas organizadas.",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(0, 12))

        linha_arquivo = ttk.Frame(frame)
        linha_arquivo.pack(fill="x", pady=(0, 12))

        ttk.Entry(
            linha_arquivo,
            textvariable=self.caminho_arquivo,
            width=64,
        ).pack(side="left", fill="x", expand=True)

        ttk.Button(
            linha_arquivo,
            text="Procurar",
            command=self.selecionar_arquivo,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            frame,
            text="Gerar Planilha Atualizada",
            command=self.executar_processamento,
        ).pack(anchor="w", pady=(0, 16))

        ttk.Label(
            frame,
            text="Status",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        self.caixa_status = tk.Text(
            frame,
            height=10,
            wrap="word",
            font=("Consolas", 10),
            state="disabled",
        )
        self.caixa_status.pack(fill="both", expand=True)

        self.atualizar_status(self.status_texto.get())

    def atualizar_status(self, mensagem):
        self.caixa_status.config(state="normal")
        self.caixa_status.delete("1.0", tk.END)
        self.caixa_status.insert(tk.END, mensagem)
        self.caixa_status.config(state="disabled")

    def selecionar_arquivo(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de calibracao",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )
        if caminho:
            self.caminho_arquivo.set(caminho)
            self.atualizar_status(f"Arquivo selecionado:\n{caminho}")

    def executar_processamento(self):
        caminho = self.caminho_arquivo.get().strip()
        if not caminho:
            messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
            return

        try:
            arquivo_saida, contagem, coluna_data = processar_planilha(caminho)
            resumo = [
                "Processamento concluido com sucesso.",
                "",
                f"Coluna usada: {coluna_data}",
                f"Aba gerada: {NOME_ABA_RESULTADO}",
                "",
                f"Arquivo gerado: {arquivo_saida}",
                "",
                "Resumo dos status:",
            ]
            for status, quantidade in contagem.items():
                resumo.append(f"- {status}: {quantidade}")

            self.atualizar_status("\n".join(resumo))
            messagebox.showinfo("Sucesso", "Planilha atualizada com sucesso.")
        except PermissionError:
            mensagem = (
                "Erro ao processar a planilha:\n"
                "Feche o arquivo no Excel e tente novamente. "
                "Se ele estiver no OneDrive, aguarde a sincronizacao terminar."
            )
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)
        except Exception as erro:
            mensagem = f"Erro ao processar a planilha:\n{erro}"
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)


def main():
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    AppCalibracao(root)
    root.mainloop()


if __name__ == "__main__":
    main()
