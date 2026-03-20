import tkinter as tk
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


NOME_ABA_PADRAO = "Cadastro"
NOME_ABA_RESULTADO = "Resultado"
TITULO_STATUS = "Status da Calibracao"

COLUNAS_DATA_CANDIDATAS = [
    "proxima calibracao",
    "data da proxima calibracao",
    "data de vencimento",
]
COLUNAS_CODIGO_CANDIDATAS = ["codigo"]
COLUNAS_PERIODICIDADE_CANDIDATAS = ["periodicidade"]
COLUNAS_NUMERO_CANDIDATAS = [
    "n\u00b0",
    "n\u00ba",
    "no",
    "n.o",
    "nro",
    "nr",
    "numero",
]

STATUS_VENCIDO = "Vencido"
STATUS_10_DIAS = "Vence em 10 dias"
STATUS_BREVE = "Vence em 30 dias"

COR_VERMELHA = "FFC7CE"
COR_AZUL = "D9EAF7"
COR_AMARELA = "FFF2CC"

PREENCHIMENTO_VERMELHO = PatternFill(
    fill_type="solid", start_color=COR_VERMELHA, end_color=COR_VERMELHA
)
PREENCHIMENTO_AZUL = PatternFill(
    fill_type="solid", start_color=COR_AZUL, end_color=COR_AZUL
)
PREENCHIMENTO_AMARELO = PatternFill(
    fill_type="solid", start_color=COR_AMARELA, end_color=COR_AMARELA
)

BORDA_FINA = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
ALINHAMENTO_CENTRO = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)


def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(char for char in texto if not unicodedata.combining(char))


def verificar_status(data, hoje):
    if data is None:
        return "Data invalida"
    if isinstance(data, datetime):
        data = data.date()
    elif isinstance(data, date):
        pass
    elif hasattr(data, "date") and not isinstance(data, str):
        data = data.date()
    else:
        return "Data invalida"

    if data < hoje:
        return STATUS_VENCIDO
    if data <= hoje + timedelta(days=10):
        return STATUS_10_DIAS
    if data <= hoje + timedelta(days=30):
        return STATUS_BREVE
    return "Em dia"


def encontrar_linha_cabecalho(ws):
    limite = min(ws.max_row, 15)
    for linha in range(1, limite + 1):
        valores = [
            normalizar_texto(ws.cell(linha, coluna).value)
            for coluna in range(1, ws.max_column + 1)
        ]
        if any(candidata in valores for candidata in COLUNAS_DATA_CANDIDATAS):
            return linha
    raise ValueError("Nao foi possivel localizar o cabecalho da planilha.")


def encontrar_coluna(ws, linha_cabecalho, candidatas, obrigatoria=True):
    for coluna in range(1, ws.max_column + 1):
        valor = normalizar_texto(ws.cell(linha_cabecalho, coluna).value)
        if valor in candidatas:
            return coluna, ws.cell(linha_cabecalho, coluna).value

    for coluna in range(1, ws.max_column + 1):
        valor = normalizar_texto(ws.cell(linha_cabecalho, coluna).value)
        if any(candidata in valor for candidata in candidatas):
            return coluna, ws.cell(linha_cabecalho, coluna).value

    if obrigatoria:
        raise ValueError("Coluna necessaria nao encontrada na planilha.")
    return None, ""


def definir_arquivo_saida(arquivo_entrada):
    base = arquivo_entrada.with_name(f"{arquivo_entrada.stem}_resultado.xlsx")
    if not base.exists():
        return base

    indice = 1
    while True:
        candidato = arquivo_entrada.with_name(
            f"{arquivo_entrada.stem}_resultado_{indice}.xlsx"
        )
        if not candidato.exists():
            return candidato
        indice += 1


def definir_arquivo_saida_alternativo(arquivo_entrada):
    pasta_local = Path(__file__).resolve().parent
    base = pasta_local / f"{arquivo_entrada.stem}_resultado.xlsx"
    if not base.exists():
        return base

    indice = 1
    while True:
        candidato = pasta_local / f"{arquivo_entrada.stem}_resultado_{indice}.xlsx"
        if not candidato.exists():
            return candidato
        indice += 1


def formatar_data(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def aplicar_cor_linha(ws, numero_linha, status):
    if status == STATUS_VENCIDO:
        preenchimento = PREENCHIMENTO_VERMELHO
    elif status == STATUS_10_DIAS:
        preenchimento = PREENCHIMENTO_AZUL
    elif status == STATUS_BREVE:
        preenchimento = PREENCHIMENTO_AMARELO
    else:
        return

    for celula in ws[numero_linha]:
        celula.fill = preenchimento


def aplicar_estilo_linha(ws, numero_linha):
    for celula in ws[numero_linha]:
        celula.border = BORDA_FINA
        celula.alignment = ALINHAMENTO_CENTRO


def ajustar_altura_linha(ws, numero_linha):
    maior_altura = 1
    for celula in ws[numero_linha]:
        valor = "" if celula.value is None else str(celula.value)
        largura = ws.column_dimensions[celula.column_letter].width or 15
        linhas_estimadas = max(1, (len(valor) // max(1, int(largura))) + 1)
        maior_altura = max(maior_altura, linhas_estimadas)

    ws.row_dimensions[numero_linha].height = max(20, maior_altura * 18)


def adicionar_resumo_e_grafico(ws, contagem):
    coluna_inicio = 10
    linha_inicio = 2
    itens_resumo = [
        ("Cor", "Significado", "Quantidade"),
        ("Vermelho", STATUS_VENCIDO, contagem.get(STATUS_VENCIDO, 0)),
        ("Azul", STATUS_10_DIAS, contagem.get(STATUS_10_DIAS, 0)),
        ("Amarelo", STATUS_BREVE, contagem.get(STATUS_BREVE, 0)),
    ]

    preenchimentos = {
        "Vermelho": PREENCHIMENTO_VERMELHO,
        "Azul": PREENCHIMENTO_AZUL,
        "Amarelo": PREENCHIMENTO_AMARELO,
    }

    for indice_linha, valores in enumerate(itens_resumo, start=linha_inicio):
        for indice_coluna, valor in enumerate(valores, start=coluna_inicio):
            celula = ws.cell(indice_linha, indice_coluna, value=valor)
            celula.border = BORDA_FINA
            celula.alignment = ALINHAMENTO_CENTRO
            if indice_linha == linha_inicio:
                celula.font = Font(bold=True)

        cor = valores[0]
        if cor in preenchimentos:
            ws.cell(indice_linha, coluna_inicio).fill = preenchimentos[cor]

    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 12

    grafico = PieChart()
    grafico.title = "Resumo da Calibracao"
    grafico.style = 10
    grafico.varyColors = False
    grafico.legend.position = "r"

    labels = Reference(
        ws,
        min_col=coluna_inicio + 1,
        min_row=linha_inicio + 1,
        max_row=linha_inicio + 3,
    )
    data = Reference(
        ws,
        min_col=coluna_inicio + 2,
        min_row=linha_inicio,
        max_row=linha_inicio + 3,
    )

    grafico.add_data(data, titles_from_data=True)
    grafico.set_categories(labels)
    grafico.height = 8
    grafico.width = 12

    grafico.dataLabels = DataLabelList()
    grafico.dataLabels.showVal = False
    grafico.dataLabels.showPercent = False
    grafico.dataLabels.showCatName = False
    grafico.dataLabels.showSerName = False
    grafico.dataLabels.showLegendKey = False

    pontos = []
    for indice, cor in enumerate([COR_VERMELHA, COR_AZUL, COR_AMARELA]):
        ponto = DataPoint(idx=indice)
        ponto.graphicalProperties = GraphicalProperties(solidFill=cor)
        pontos.append(ponto)

    grafico.series[0].data_points = pontos
    ws.add_chart(grafico, "J7")


def criar_planilha_resultado(registros, arquivo_saida, contagem):
    wb = Workbook()
    ws = wb.active
    ws.title = NOME_ABA_RESULTADO

    cabecalhos = [
        "Codigo",
        "N\u00b0",
        "Periodicidade",
        "Descricao",
        "Setor Utilizado",
        "Proxima Calibracao",
        TITULO_STATUS,
    ]
    ws.append(cabecalhos)

    for celula in ws[1]:
        celula.font = Font(bold=True)
    aplicar_estilo_linha(ws, 1)

    larguras = {
        "A": 18,
        "B": 12,
        "C": 20,
        "D": 70,
        "E": 28,
        "F": 18,
        "G": 22,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    for registro in registros:
        ws.append(registro)
        aplicar_estilo_linha(ws, ws.max_row)
        aplicar_cor_linha(ws, ws.max_row, registro[-1])
        ajustar_altura_linha(ws, ws.max_row)

    ajustar_altura_linha(ws, 1)
    adicionar_resumo_e_grafico(ws, contagem)
    ws.freeze_panes = "A2"
    wb.save(arquivo_saida)
    wb.close()


def processar_planilha(caminho_arquivo):
    arquivo_entrada = Path(caminho_arquivo)
    if not arquivo_entrada.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {arquivo_entrada}")

    arquivo_saida = definir_arquivo_saida(arquivo_entrada)
    wb_valores = load_workbook(arquivo_entrada, data_only=True)

    if NOME_ABA_PADRAO not in wb_valores.sheetnames:
        raise ValueError(f"A aba '{NOME_ABA_PADRAO}' nao foi encontrada.")

    ws_valores = wb_valores[NOME_ABA_PADRAO]
    linha_cabecalho = encontrar_linha_cabecalho(ws_valores)
    coluna_data, nome_coluna_data = encontrar_coluna(
        ws_valores, linha_cabecalho, COLUNAS_DATA_CANDIDATAS
    )
    coluna_codigo, _ = encontrar_coluna(
        ws_valores, linha_cabecalho, COLUNAS_CODIGO_CANDIDATAS
    )
    coluna_periodicidade, _ = encontrar_coluna(
        ws_valores,
        linha_cabecalho,
        COLUNAS_PERIODICIDADE_CANDIDATAS,
        obrigatoria=False,
    )
    coluna_numero, _ = encontrar_coluna(
        ws_valores,
        linha_cabecalho,
        COLUNAS_NUMERO_CANDIDATAS,
        obrigatoria=False,
    )

    hoje = datetime.today().date()
    contagem = {STATUS_VENCIDO: 0, STATUS_10_DIAS: 0, STATUS_BREVE: 0}
    registros = []

    for linha in range(linha_cabecalho + 1, ws_valores.max_row + 1):
        codigo = ws_valores.cell(linha, coluna_codigo).value
        proxima_calibracao = ws_valores.cell(linha, coluna_data).value

        if codigo in (None, ""):
            continue
        if proxima_calibracao in (None, ""):
            continue

        status = verificar_status(proxima_calibracao, hoje)
        if status not in (STATUS_VENCIDO, STATUS_10_DIAS, STATUS_BREVE):
            continue

        contagem[status] = contagem.get(status, 0) + 1

        periodicidade = ""
        if coluna_periodicidade:
            periodicidade = ws_valores.cell(linha, coluna_periodicidade).value or ""

        numero = ""
        if coluna_numero:
            numero = ws_valores.cell(linha, coluna_numero).value or ""

        registros.append(
            [
                codigo,
                numero,
                periodicidade,
                ws_valores.cell(linha, 7).value or "",
                ws_valores.cell(linha, 9).value or "",
                formatar_data(proxima_calibracao),
                status,
            ]
        )

    wb_valores.close()

    try:
        criar_planilha_resultado(registros, arquivo_saida, contagem)
    except PermissionError:
        arquivo_saida = definir_arquivo_saida_alternativo(arquivo_entrada)
        criar_planilha_resultado(registros, arquivo_saida, contagem)

    contagem = {chave: valor for chave, valor in contagem.items() if valor > 0}
    return arquivo_saida, contagem, nome_coluna_data


class AppCalibracao:
    def __init__(self, root):
        self.root = root
        self.root.title("Controle de Calibracao")
        self.root.geometry("700x340")
        self.root.resizable(False, False)

        self.caminho_arquivo = tk.StringVar()
        self.status_texto = tk.StringVar(
            value=(
                "Selecione a planilha. O programa vai usar apenas a aba Cadastro e "
                "gerar um Excel apenas com itens vencidos ou que vencem em breve."
            )
        )
        self.montar_tela()

    def montar_tela(self):
        frame = ttk.Frame(self.root, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(
            frame,
            text="Controle de Calibracao",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", pady=(0, 12))

        ttk.Label(
            frame,
            text="Itens vencidos ficam em vermelho, 10 dias em azul e 30 dias em amarelo.",
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(0, 12))

        linha_arquivo = ttk.Frame(frame)
        linha_arquivo.pack(fill="x", pady=(0, 12))

        ttk.Entry(
            linha_arquivo,
            textvariable=self.caminho_arquivo,
            width=64,
        ).pack(side="left", fill="x", expand=True)

        ttk.Button(
            linha_arquivo,
            text="Procurar",
            command=self.selecionar_arquivo,
        ).pack(side="left", padx=(8, 0))

        ttk.Button(
            frame,
            text="Gerar Planilha Atualizada",
            command=self.executar_processamento,
        ).pack(anchor="w", pady=(0, 16))

        ttk.Label(
            frame,
            text="Status",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w")

        self.caixa_status = tk.Text(
            frame,
            height=10,
            wrap="word",
            font=("Consolas", 10),
            state="disabled",
        )
        self.caixa_status.pack(fill="both", expand=True)

        self.atualizar_status(self.status_texto.get())

    def atualizar_status(self, mensagem):
        self.caixa_status.config(state="normal")
        self.caixa_status.delete("1.0", tk.END)
        self.caixa_status.insert(tk.END, mensagem)
        self.caixa_status.config(state="disabled")

    def selecionar_arquivo(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de calibracao",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
        )
        if caminho:
            self.caminho_arquivo.set(caminho)
            self.atualizar_status(f"Arquivo selecionado:\n{caminho}")

    def executar_processamento(self):
        caminho = self.caminho_arquivo.get().strip()
        if not caminho:
            messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
            return

        try:
            arquivo_saida, contagem, coluna_data = processar_planilha(caminho)
            resumo = [
                "Processamento concluido com sucesso.",
                "",
                f"Aba usada: {NOME_ABA_PADRAO}",
                f"Coluna usada: {coluna_data}",
                f"Aba gerada: {NOME_ABA_RESULTADO}",
                "",
                f"Arquivo gerado: {arquivo_saida}",
                "",
                "Resumo dos status:",
            ]
            if contagem:
                for status, quantidade in contagem.items():
                    resumo.append(f"- {status}: {quantidade}")
            else:
                resumo.append("- Nenhum item vencido ou a vencer em breve.")

            self.atualizar_status("\n".join(resumo))
            messagebox.showinfo("Sucesso", "Planilha atualizada com sucesso.")
        except PermissionError:
            mensagem = (
                "Erro ao processar a planilha:\n"
                "Feche o arquivo no Excel e tente novamente. "
                "Se ele estiver no OneDrive, aguarde a sincronizacao terminar."
            )
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)
        except Exception as erro:
            mensagem = f"Erro ao processar a planilha:\n{erro}"
            self.atualizar_status(mensagem)
            messagebox.showerror("Erro", mensagem)


def main():
    root = tk.Tk()
    style = ttk.Style()
    if "vista" in style.theme_names():
        style.theme_use("vista")
    AppCalibracao(root)
    root.mainloop()


if __name__ == "__main__":
    main()
